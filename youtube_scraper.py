#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
YouTube 채널 최근 한 달 영상 정보 추출기
=============================================
API 키 없이 yt-dlp를 사용해 채널 영상을 긁어오고
제목 / 조회수 / 업로드 날짜 / 썸네일 URL을 엑셀로 저장합니다.

설치:
    pip install yt-dlp openpyxl

사용법:
    1. 아래 CHANNEL_URLS 에 원하는 채널 주소를 추가
    2. python youtube_scraper.py 실행
    3. 바탕화면에 엑셀 파일 생성 확인
"""

import os
import sys
import time
from datetime import datetime, timedelta

try:
    import yt_dlp
except ImportError:
    print("❌ yt-dlp가 설치되어 있지 않습니다.")
    print("   터미널에서 실행: pip install yt-dlp")
    sys.exit(1)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("❌ openpyxl이 설치되어 있지 않습니다.")
    print("   터미널에서 실행: pip install openpyxl")
    sys.exit(1)


# ============================================================
# ✅ 여기에 크롤링할 채널 URL을 입력하세요
#    다양한 형식 모두 지원:
#    - https://www.youtube.com/@핸들명
#    - https://www.youtube.com/channel/UC...
#    - https://www.youtube.com/c/채널명
# ============================================================
CHANNEL_URLS = [
    "https://www.youtube.com/@MrBeast",
    "https://www.youtube.com/@mkbhd",
    # 원하는 채널을 계속 추가하세요...
]

# ============================================================
# ⚙️ 세부 설정
# ============================================================
DAYS_BACK = 30          # 최근 며칠치 영상을 가져올지 (기본 30일)
MAX_VIDEOS = 50         # 채널당 최대 스캔할 영상 수 (많을수록 느림)
                        # → 영상을 자주 올리는 채널은 늘려주세요
RETRY_COUNT = 3         # 실패 시 재시도 횟수
RETRY_DELAY = 3         # 재시도 간격 (초)


# ============================================================
# 📁 저장 경로 (Windows 바탕화면)
# ============================================================
desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
OUTPUT_FILENAME = f"유튜브_영상목록_{timestamp}.xlsx"
OUTPUT_PATH = os.path.join(desktop_path, OUTPUT_FILENAME)


# ============================================================
# 유튜브 데이터 수집 함수
# ============================================================

def fetch_video_list(channel_url: str) -> tuple[str, list[dict]]:
    """
    채널에서 영상 목록을 빠르게 가져옵니다 (extract_flat 사용).
    반환: (채널명, [{'id', 'title', 'upload_date', 'view_count', 'thumbnail'}, ...])
    """
    url = channel_url.rstrip("/") + "/videos"

    ydl_opts = {
        "extract_flat": "in_playlist",
        "playlistend": MAX_VIDEOS,
        "quiet": True,
        "no_warnings": True,
        "ignoreerrors": True,
    }

    with yt_dlp.YoutubeDL(ydl_opts) as ydl:
        info = ydl.extract_info(url, download=False)

    if not info:
        return channel_url, []

    channel_name = (
        info.get("channel")
        or info.get("uploader")
        or info.get("title")
        or channel_url
    )
    entries = info.get("entries") or []

    videos = []
    for entry in entries:
        if not entry:
            continue
        videos.append({
            "id": entry.get("id", ""),
            "title": entry.get("title", "(제목 없음)"),
            "upload_date": entry.get("upload_date"),   # "YYYYMMDD" 문자열 or None
            "view_count": entry.get("view_count"),      # None일 수 있음
            "thumbnail": entry.get("thumbnail", ""),
        })

    return channel_name, videos


def fetch_video_detail(video_id: str) -> dict:
    """
    개별 영상 ID로 상세 정보(조회수, 썸네일 등)를 가져옵니다.
    """
    ydl_opts = {
        "quiet": True,
        "no_warnings": True,
        "skip_download": True,
        "ignoreerrors": True,
    }
    url = f"https://www.youtube.com/watch?v={video_id}"
    with yt_dlp.YoutubeDL(ydl_opts) as ydl:
        info = ydl.extract_info(url, download=False)
    return info or {}


def get_channel_videos(channel_url: str) -> list[dict]:
    """
    채널 URL에서 최근 DAYS_BACK일 내 영상 정보를 수집해 반환합니다.
    """
    cutoff = datetime.now() - timedelta(days=DAYS_BACK)
    print(f"\n  📡 [{channel_url}] 영상 목록 수집 중...")

    # ── 1단계: 빠른 목록 수집 ──────────────────────────────
    channel_name = channel_url
    raw_list = []

    for attempt in range(1, RETRY_COUNT + 1):
        try:
            channel_name, raw_list = fetch_video_list(channel_url)
            break
        except Exception as e:
            if attempt < RETRY_COUNT:
                print(f"  ⚠️ 시도 {attempt}/{RETRY_COUNT} 실패, {RETRY_DELAY}초 후 재시도... ({e})")
                time.sleep(RETRY_DELAY)
            else:
                print(f"  ❌ 채널 정보 가져오기 최종 실패: {e}")
                return []

    print(f"  ✔ 채널: {channel_name} | 스캔 영상 수: {len(raw_list)}개")

    # ── 2단계: 날짜 필터링 ─────────────────────────────────
    in_range: list[dict] = []
    for v in raw_list:
        if v["upload_date"]:
            upload_dt = datetime.strptime(v["upload_date"], "%Y%m%d")
            if upload_dt >= cutoff:
                v["_upload_dt"] = upload_dt
                in_range.append(v)
        else:
            # 날짜를 알 수 없는 경우 → 개별 조회 대상에 포함
            v["_upload_dt"] = None
            in_range.append(v)

    print(f"  ✔ 최근 {DAYS_BACK}일 내 영상(후보): {len(in_range)}개")

    # ── 3단계: 조회수 누락 영상 개별 조회 ─────────────────
    needs_detail = [v for v in in_range if v["view_count"] is None and v["id"]]
    if needs_detail:
        print(f"  🔍 조회수 누락 → 개별 조회 ({len(needs_detail)}개)...")
        for i, v in enumerate(needs_detail, 1):
            try:
                detail = fetch_video_detail(v["id"])
                v["view_count"] = detail.get("view_count", 0)
                if not v["thumbnail"]:
                    v["thumbnail"] = detail.get("thumbnail", "")
                if v["_upload_dt"] is None:
                    ds = detail.get("upload_date", "")
                    if ds:
                        v["_upload_dt"] = datetime.strptime(ds, "%Y%m%d")
                sys.stdout.write(f"\r    진행: {i}/{len(needs_detail)}")
                sys.stdout.flush()
            except Exception as e:
                pass  # 실패 시 해당 영상은 0으로 처리
        print()  # 줄바꿈

    # ── 4단계: 최종 날짜 재필터 & 정리 ───────────────────
    result = []
    for v in in_range:
        dt = v.get("_upload_dt")
        # 날짜를 끝내 알 수 없으면 제외
        if dt is None:
            continue
        if dt < cutoff:
            continue

        result.append({
            "channel": channel_name,
            "title": v["title"],
            "view_count": v["view_count"] if v["view_count"] is not None else 0,
            "upload_date": dt.strftime("%Y-%m-%d"),
            "thumbnail": v["thumbnail"],
            "video_url": f"https://www.youtube.com/watch?v={v['id']}",
        })

    print(f"  ✅ 최종 수집: {len(result)}개")
    return result


# ============================================================
# 엑셀 저장 함수
# ============================================================

def save_excel(all_videos: list[dict], output_path: str) -> None:
    """수집된 영상 데이터를 보기 좋은 엑셀 파일로 저장합니다."""

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "YouTube 영상 목록"

    # ── 스타일 정의 ────────────────────────────────────────
    HEADER_BG   = "2563EB"   # 파란색
    ROW_ODD_BG  = "FFFFFF"   # 흰색
    ROW_EVEN_BG = "EFF6FF"   # 연한 파란색

    header_font = Font(bold=True, color="FFFFFF", size=11, name="맑은 고딕")
    data_font   = Font(size=10, name="맑은 고딕")
    header_fill = PatternFill("solid", fgColor=HEADER_BG)
    even_fill   = PatternFill("solid", fgColor=ROW_EVEN_BG)

    center = Alignment(horizontal="center", vertical="center", wrap_text=False)
    left   = Alignment(horizontal="left",   vertical="center", wrap_text=False)
    left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # ── 헤더 ───────────────────────────────────────────────
    headers = [
        ("채널명",      20, center),
        ("영상 제목",   55, left_wrap),
        ("조회수",      14, center),
        ("업로드 날짜", 14, center),
        ("썸네일 URL",  55, left),
        ("영상 URL",    45, left),
    ]

    for col_idx, (label, width, _align) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_fill and header_font
        cell.fill = header_fill
        cell.alignment = center
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 28

    # ── 데이터 행 ──────────────────────────────────────────
    for row_idx, video in enumerate(all_videos, start=2):
        is_even = (row_idx % 2 == 0)
        row_fill = even_fill if is_even else None

        values = [
            video["channel"],
            video["title"],
            video["view_count"],
            video["upload_date"],
            video["thumbnail"],
            video["video_url"],
        ]
        aligns = [center, left_wrap, center, center, left, left]

        for col_idx, (val, align) in enumerate(zip(values, aligns), 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = data_font
            cell.alignment = align
            if row_fill:
                cell.fill = row_fill

        # 조회수 숫자 포맷 (#,##0)
        ws.cell(row=row_idx, column=3).number_format = "#,##0"

        ws.row_dimensions[row_idx].height = 20

    # ── 틀 고정 & 자동 필터 ────────────────────────────────
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(all_videos) + 1}"

    # ── 요약 시트 ──────────────────────────────────────────
    ws_summary = wb.create_sheet("채널별 요약")
    ws_summary.column_dimensions["A"].width = 35
    ws_summary.column_dimensions["B"].width = 15

    sum_headers = ["채널명", "영상 수"]
    for col_idx, h in enumerate(sum_headers, 1):
        cell = ws_summary.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    from collections import Counter
    channel_counts = Counter(v["channel"] for v in all_videos)
    for row_idx, (ch, cnt) in enumerate(
        sorted(channel_counts.items(), key=lambda x: -x[1]), start=2
    ):
        ws_summary.cell(row=row_idx, column=1, value=ch).alignment = left
        ws_summary.cell(row=row_idx, column=2, value=cnt).alignment = center

    wb.save(output_path)


# ============================================================
# 메인
# ============================================================

def main():
    print("=" * 60)
    print("🎬  YouTube 채널 영상 정보 추출기")
    print(f"📅  최근 {DAYS_BACK}일간 영상 수집")
    print(f"📁  저장 경로: {OUTPUT_PATH}")
    print("=" * 60)

    if not CHANNEL_URLS:
        print("❌ CHANNEL_URLS 목록이 비어 있습니다. 스크립트 상단에 채널 URL을 추가하세요.")
        sys.exit(1)

    all_videos: list[dict] = []

    for idx, url in enumerate(CHANNEL_URLS, 1):
        print(f"\n[{idx}/{len(CHANNEL_URLS)}] 처리 중...")
        try:
            videos = get_channel_videos(url)
            all_videos.extend(videos)
        except KeyboardInterrupt:
            print("\n⏹  사용자 중단 (지금까지 수집된 데이터를 저장합니다)")
            break
        except Exception as e:
            print(f"  ❌ 예상치 못한 오류: {e}")

    if not all_videos:
        print("\n⚠️  수집된 영상이 없습니다. 채널 URL을 확인해 주세요.")
        sys.exit(0)

    # 최신 날짜 기준 정렬
    all_videos.sort(key=lambda x: x["upload_date"], reverse=True)

    print(f"\n{'=' * 60}")
    print(f"📊  총 {len(all_videos)}개 영상 수집 완료")
    print(f"💾  엑셀 저장 중...")

    try:
        save_excel(all_videos, OUTPUT_PATH)
    except PermissionError:
        # 파일이 열려 있을 경우 다른 이름으로 저장
        alt_path = OUTPUT_PATH.replace(".xlsx", "_new.xlsx")
        save_excel(all_videos, alt_path)
        print(f"\n⚠️  기존 파일이 열려 있어 다른 이름으로 저장했습니다: {alt_path}")
    else:
        print(f"✅  저장 완료 → 바탕화면 [{OUTPUT_FILENAME}]")

    print("=" * 60)


if __name__ == "__main__":
    main()
