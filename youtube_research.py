#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
YouTube 트렌드 리서처
─────────────────────
키워드 또는 채널 URL로 유튜브 영상을 수집하고
기여도 · 성과도 · 노출확률 등 지표를 분석하는 Streamlit 웹앱
"""

import io
import time
from datetime import datetime, timedelta

import pandas as pd
import streamlit as st

try:
    import yt_dlp
except ImportError:
    st.error("yt-dlp 패키지가 없습니다. 터미널에서 `pip install yt-dlp` 실행 후 새로고침하세요.")
    st.stop()

# ════════════════════════════════════════════════════════════
#  페이지 기본 설정
# ════════════════════════════════════════════════════════════
st.set_page_config(
    page_title="YouTube 트렌드 리서처",
    page_icon="🎬",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── 커스텀 CSS ──────────────────────────────────────────────
st.markdown("""
<style>
/* 헤더 */
.main-header {
    background: linear-gradient(135deg, #1e3a8a 0%, #3b82f6 100%);
    padding: 1.5rem 2rem;
    border-radius: 12px;
    color: white;
    margin-bottom: 1.5rem;
}
.main-header h1 { margin: 0; font-size: 1.8rem; }
.main-header p  { margin: 0.3rem 0 0; opacity: 0.85; font-size: 0.9rem; }

/* 지표 뱃지 */
.badge-vg  { color: #15803d; font-weight: 700; }
.badge-g   { color: #16a34a; font-weight: 600; }
.badge-n   { color: #d97706; font-weight: 600; }
.badge-l   { color: #dc2626; font-weight: 600; }

/* 메트릭 카드 */
div[data-testid="metric-container"] {
    background: #f8fafc;
    border: 1px solid #e2e8f0;
    border-radius: 10px;
    padding: 0.8rem 1rem;
}

/* 검색 박스 강조 */
div[data-testid="stTextInput"] input {
    border-radius: 8px !important;
}
</style>
""", unsafe_allow_html=True)


# ════════════════════════════════════════════════════════════
#  유틸리티 함수
# ════════════════════════════════════════════════════════════

def fmt_num(n: int | float | None, unit: str = "") -> str:
    """숫자를 한국식 단위로 포맷팅"""
    if n is None or n == 0:
        return "-"
    n = int(n)
    if n >= 100_000_000:
        return f"{n / 100_000_000:.1f}억{unit}"
    if n >= 10_000:
        return f"{n / 10_000:.1f}만{unit}"
    if n >= 1_000:
        return f"{n / 1_000:.1f}천{unit}"
    return f"{n:,}{unit}"


def perf_grade(ratio: float | None) -> str:
    """성과도: 결과 내 중앙값 대비 조회수 배율"""
    if ratio is None:
        return "─"
    if ratio >= 3.0:
        return "🔥 Very Good"
    if ratio >= 1.5:
        return "✅ Good"
    if ratio >= 0.5:
        return "📊 Normal"
    return "📉 Low"


def contrib_grade(rate: float | None) -> str:
    """기여도: 구독자 대비 조회수 % (채널 파급력)"""
    if rate is None:
        return "─"
    if rate >= 20:
        return "🔥 Very Good"
    if rate >= 8:
        return "✅ Good"
    if rate >= 2:
        return "📊 Normal"
    return "📉 Low"


def exposure_grade(vpd: float | None) -> str:
    """노출확률: 일평균 조회수 기반 바이럴 속도"""
    if vpd is None:
        return "─"
    if vpd >= 50_000:
        return "🔥 Very High"
    if vpd >= 10_000:
        return "✅ High"
    if vpd >= 1_000:
        return "📊 Medium"
    return "📉 Low"


# ════════════════════════════════════════════════════════════
#  yt-dlp 데이터 수집 함수
# ════════════════════════════════════════════════════════════

@st.cache_data(ttl=1800, show_spinner=False)
def search_by_keyword(keyword: str, max_results: int = 50) -> list[dict]:
    """키워드로 YouTube 영상 검색 (ytsearch)"""
    ydl_opts = {
        "extract_flat": True,
        "quiet": True,
        "no_warnings": True,
        "ignoreerrors": True,
    }
    try:
        with yt_dlp.YoutubeDL(ydl_opts) as ydl:
            info = ydl.extract_info(f"ytsearch{max_results}:{keyword}", download=False)
        return [e for e in (info.get("entries") or []) if e] if info else []
    except Exception as e:
        st.error(f"키워드 검색 오류: {e}")
        return []


@st.cache_data(ttl=1800, show_spinner=False)
def search_by_channel(channel_url: str, max_videos: int, days_back: int) -> tuple[str, list[dict]]:
    """채널 URL에서 영상 목록 수집 (여러 URL 형식 자동 시도)"""
    base = channel_url.rstrip("/")
    urls_to_try = [base + "/videos", base, base + "/streams"]

    ydl_opts = {
        "extract_flat": "in_playlist",
        "playlistend": max_videos,
        "quiet": True,
        "no_warnings": True,
        "ignoreerrors": True,
    }

    info = None
    for url in urls_to_try:
        try:
            with yt_dlp.YoutubeDL(ydl_opts) as ydl:
                result = ydl.extract_info(url, download=False)
            if result and result.get("entries"):
                info = result
                break
        except Exception:
            continue

    if not info:
        return "", []

    channel_name = (
        info.get("channel") or info.get("uploader") or info.get("title") or channel_url
    )

    cutoff = datetime.now() - timedelta(days=days_back)
    entries = []
    for e in (info.get("entries") or []):
        if not e:
            continue
        ds = e.get("upload_date")
        if ds:
            try:
                if datetime.strptime(ds, "%Y%m%d") < cutoff:
                    continue
            except ValueError:
                pass
        entries.append(e)

    return channel_name, entries


# ════════════════════════════════════════════════════════════
#  엔트리 → DataFrame 변환 + 지표 계산
# ════════════════════════════════════════════════════════════

def build_dataframe(entries: list[dict], source_label: str = "") -> pd.DataFrame:
    today = datetime.now()
    rows = []

    for e in entries:
        if not e:
            continue

        vid_id = e.get("id", "")

        # ── 날짜 파싱 ──────────────────────────────────────
        date_str = e.get("upload_date")
        upload_dt, days_old = None, None
        if date_str:
            try:
                upload_dt = datetime.strptime(date_str, "%Y%m%d")
                days_old = max((today - upload_dt).days, 1)
            except ValueError:
                pass

        # ── 조회수 / 구독자 ────────────────────────────────
        view_count = e.get("view_count")
        sub_count = (
            e.get("channel_follower_count")
            or e.get("subscriber_count")
        )

        # ── 파생 지표 계산 ─────────────────────────────────
        vpd = round(view_count / days_old, 1) if (view_count and days_old) else None
        contrib_rate = (
            round(view_count / sub_count * 100, 2)
            if (view_count and sub_count and sub_count > 0)
            else None
        )

        rows.append({
            # 내부 계산용 (비표시)
            "_id": vid_id,
            "_view_raw": view_count or 0,
            "_vpd": vpd,
            "_contrib_rate": contrib_rate,
            "_source": source_label,
            # 표시용
            "썸네일": e.get("thumbnail", ""),
            "제목": e.get("title", "(제목없음)"),
            "채널": e.get("channel") or e.get("uploader") or "-",
            "조회수_raw": view_count,
            "구독자_raw": sub_count,
            "게시일": upload_dt.strftime("%Y-%m-%d") if upload_dt else "-",
            "링크": f"https://www.youtube.com/watch?v={vid_id}" if vid_id else "",
        })

    if not rows:
        return pd.DataFrame()

    df = pd.DataFrame(rows)

    # ── 성과도: 결과 내 중앙값 대비 ───────────────────────
    median_v = df["_view_raw"].median()
    df["_perf_ratio"] = df["_view_raw"].apply(
        lambda v: round(v / median_v, 2) if median_v > 0 else None
    )

    # ── 표시용 컬럼 생성 ───────────────────────────────────
    df["조회수"] = df["조회수_raw"].apply(fmt_num)
    df["구독자"] = df["구독자_raw"].apply(fmt_num)
    df["기여도"] = df["_contrib_rate"].apply(contrib_grade)
    df["성과도"] = df["_perf_ratio"].apply(perf_grade)
    df["노출확률"] = df["_vpd"].apply(exposure_grade)
    df["일평균조회"] = df["_vpd"].apply(lambda x: fmt_num(x, "회/일") if x else "-")

    return df


# ════════════════════════════════════════════════════════════
#  엑셀 변환
# ════════════════════════════════════════════════════════════

def to_excel_bytes(df: pd.DataFrame) -> bytes:
    cols = {
        "채널": "채널명",
        "제목": "영상 제목",
        "_view_raw": "조회수",
        "구독자_raw": "구독자 수",
        "기여도": "기여도",
        "성과도": "성과도",
        "노출확률": "노출확률",
        "일평균조회": "일평균 조회수",
        "게시일": "게시일",
        "링크": "영상 URL",
        "썸네일": "썸네일 URL",
    }
    export_df = df[[c for c in cols if c in df.columns]].rename(columns=cols)

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        export_df.to_excel(writer, index=False, sheet_name="YouTube 분석")

        # 열 너비 자동 조정
        ws = writer.sheets["YouTube 분석"]
        widths = [20, 55, 12, 12, 14, 14, 14, 16, 12, 45, 55]
        from openpyxl.utils import get_column_letter
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # 헤더 스타일
        from openpyxl.styles import Font, PatternFill, Alignment
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF", name="맑은 고딕")
            cell.fill = PatternFill("solid", fgColor="2563EB")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.freeze_panes = "A2"

    buf.seek(0)
    return buf.read()


# ════════════════════════════════════════════════════════════
#  사이드바 필터
# ════════════════════════════════════════════════════════════

with st.sidebar:
    st.markdown("## ⚙️ 검색 설정")

    max_results = st.slider("최대 수집 영상 수", 10, 100, 50, 10,
                            help="많을수록 시간이 걸립니다")
    days_back = st.slider("채널 검색 기간 (최근 N일)", 7, 365, 30, 7,
                          help="채널 URL 검색 시 적용됩니다")

    st.divider()
    st.markdown("## 🔽 결과 필터")

    sort_opt = st.selectbox(
        "정렬 기준",
        ["조회수 높은순", "최신순", "일평균 조회수 높은순", "관련도순 (검색 기본)"],
    )
    min_views = st.number_input("최소 조회수", min_value=0, value=0, step=10_000,
                                format="%d")
    perf_filter = st.multiselect(
        "성과도 필터",
        ["🔥 Very Good", "✅ Good", "📊 Normal", "📉 Low"],
        default=[],
        help="선택한 성과도 등급만 표시합니다",
    )
    exposure_filter = st.multiselect(
        "노출확률 필터",
        ["🔥 Very High", "✅ High", "📊 Medium", "📉 Low"],
        default=[],
    )

    st.divider()
    st.caption("ℹ️ 지표 설명")
    with st.expander("기여도 · 성과도 · 노출확률이란?"):
        st.markdown("""
| 지표 | 계산 방식 |
|------|----------|
| **기여도** | 조회수 ÷ 구독자 수 × 100% |
| **성과도** | 조회수 ÷ 결과 내 중앙값 |
| **노출확률** | 일평균 조회수 (바이럴 속도) |

- **구독자 수**는 채널 URL 검색 시 더 정확합니다
- 키워드 검색은 구독자 정보가 없을 수 있습니다
        """)


# ════════════════════════════════════════════════════════════
#  메인 UI
# ════════════════════════════════════════════════════════════

st.markdown("""
<div class="main-header">
  <h1>🎬 YouTube 트렌드 리서처</h1>
  <p>키워드 또는 채널 URL로 유튜브 영상을 분석하고 기여도 · 성과도 · 노출확률을 한눈에 확인하세요</p>
</div>
""", unsafe_allow_html=True)

# 검색 입력 영역
col_kw, col_ch = st.columns([1, 1], gap="medium")

with col_kw:
    keyword = st.text_input(
        "🔍 키워드 검색",
        placeholder="예: 병원 마케팅, 유튜브 알고리즘 ...",
        help="유튜브 전체에서 관련 영상을 검색합니다",
    )

with col_ch:
    channel_url = st.text_input(
        "📺 채널 URL (선택)",
        placeholder="https://www.youtube.com/@채널명",
        help="특정 채널의 최근 영상만 수집합니다",
    )

btn_col, _ = st.columns([1, 4])
with btn_col:
    search_btn = st.button("🔍 검색 시작", type="primary", use_container_width=True)


# ════════════════════════════════════════════════════════════
#  검색 실행
# ════════════════════════════════════════════════════════════

if search_btn:
    if not keyword and not channel_url:
        st.warning("키워드 또는 채널 URL을 입력해 주세요.")
        st.stop()

    all_entries: list[dict] = []
    labels: list[str] = []

    # ── 키워드 검색 ────────────────────────────────────────
    if keyword:
        with st.spinner(f"🔍 '{keyword}' 검색 중..."):
            kw_entries = search_by_keyword(keyword, max_results)
        if kw_entries:
            all_entries.extend(kw_entries)
            labels.append(f"🔍 키워드: **{keyword}**")
            st.success(f"키워드 검색 완료 — {len(kw_entries)}개 발견")
        else:
            st.warning("키워드 검색 결과가 없습니다.")

    # ── 채널 검색 ──────────────────────────────────────────
    if channel_url:
        with st.spinner("📡 채널 영상 수집 중..."):
            ch_name, ch_entries = search_by_channel(channel_url, max_results, days_back)
        if ch_entries:
            all_entries.extend(ch_entries)
            labels.append(f"📺 채널: **{ch_name}** (최근 {days_back}일)")
            st.success(f"채널 수집 완료 — {len(ch_entries)}개 발견")
        else:
            st.error(
                "채널 영상을 가져오지 못했습니다. "
                "URL을 확인해 주세요. (예: https://www.youtube.com/@채널핸들)"
            )

    if not all_entries:
        st.stop()

    # ── DataFrame 생성 ─────────────────────────────────────
    df = build_dataframe(all_entries)

    if df.empty:
        st.warning("수집된 영상이 없습니다.")
        st.stop()

    # ── 필터 적용 ──────────────────────────────────────────
    if min_views > 0:
        df = df[df["_view_raw"] >= min_views]
    if perf_filter:
        df = df[df["성과도"].isin(perf_filter)]
    if exposure_filter:
        df = df[df["노출확률"].isin(exposure_filter)]

    if df.empty:
        st.warning("필터 조건에 맞는 영상이 없습니다. 사이드바 필터를 완화해 보세요.")
        st.stop()

    # ── 정렬 ───────────────────────────────────────────────
    if sort_opt == "조회수 높은순":
        df = df.sort_values("_view_raw", ascending=False)
    elif sort_opt == "최신순":
        df = df.sort_values("게시일", ascending=False)
    elif sort_opt == "일평균 조회수 높은순":
        df = df.sort_values("_vpd", ascending=False, na_position="last")
    # 관련도순은 원래 순서 유지

    df = df.reset_index(drop=True)

    # ── 요약 지표 ──────────────────────────────────────────
    st.divider()
    st.markdown(" / ".join(labels))

    m1, m2, m3, m4, m5 = st.columns(5)
    m1.metric("📊 총 영상 수", f"{len(df):,}개")
    m2.metric("👁️ 평균 조회수",
              fmt_num(int(df["_view_raw"].mean())) if df["_view_raw"].sum() > 0 else "-")
    m3.metric("🔥 최고 조회수",
              fmt_num(int(df["_view_raw"].max())) if df["_view_raw"].sum() > 0 else "-")
    m4.metric("📅 가장 최신",
              df[df["게시일"] != "-"]["게시일"].max() if (df["게시일"] != "-").any() else "-")
    good_count = df["성과도"].str.contains("Good|Very").sum()
    m5.metric("✅ Good 이상", f"{good_count}개 ({int(good_count/len(df)*100)}%)")

    st.divider()

    # ── 결과 테이블 ────────────────────────────────────────
    display_cols = [
        "썸네일", "제목", "채널",
        "조회수", "구독자",
        "기여도", "성과도", "노출확률",
        "일평균조회", "게시일", "링크",
    ]
    display_df = df[display_cols].copy()

    st.data_editor(
        display_df,
        column_config={
            "썸네일": st.column_config.ImageColumn("썸네일", width="small"),
            "제목": st.column_config.TextColumn("제목", width="large"),
            "채널": st.column_config.TextColumn("채널", width="medium"),
            "조회수": st.column_config.TextColumn("조회수", width="small"),
            "구독자": st.column_config.TextColumn("구독자", width="small"),
            "기여도": st.column_config.TextColumn("기여도", width="medium"),
            "성과도": st.column_config.TextColumn("성과도", width="medium"),
            "노출확률": st.column_config.TextColumn("노출확률", width="medium"),
            "일평균조회": st.column_config.TextColumn("일평균조회", width="medium"),
            "게시일": st.column_config.TextColumn("게시일", width="small"),
            "링크": st.column_config.LinkColumn("바로가기", display_text="▶ 보기"),
        },
        hide_index=True,
        use_container_width=True,
        disabled=True,
        height=min(600, 80 + len(df) * 55),
        key="result_table",
    )

    # ── 엑셀 다운로드 ──────────────────────────────────────
    st.divider()
    dl_col, info_col = st.columns([1, 3])
    with dl_col:
        fname = f"youtube_분석_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        st.download_button(
            "💾 엑셀 다운로드",
            data=to_excel_bytes(df),
            file_name=fname,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    with info_col:
        st.caption(
            f"총 {len(df)}개 영상 데이터가 포함됩니다. "
            "구독자 수는 키워드 검색 시 제공되지 않을 수 있습니다."
        )

# ── 초기 화면 ──────────────────────────────────────────────
else:
    st.info("검색어 또는 채널 URL을 입력하고 **검색 시작** 버튼을 눌러주세요.")

    st.markdown("### 💡 사용 예시")
    ex1, ex2, ex3 = st.columns(3)

    with ex1:
        st.markdown("""
**📌 키워드만 검색**
```
키워드: 병원 마케팅
```
유튜브 전체에서 관련 영상 상위 N개 수집
""")

    with ex2:
        st.markdown("""
**📌 채널만 검색**
```
채널 URL: youtube.com/@채널명
```
해당 채널의 최근 N일 영상만 수집
""")

    with ex3:
        st.markdown("""
**📌 둘 다 입력**
```
키워드: 마케팅
채널 URL: youtube.com/@채널
```
키워드 결과 + 채널 영상 합쳐서 표시
""")
